from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import load_workbook

from copy import copy

import openpyxl

from PIL import Image


def style_range(ws, cell_range, border=Border(), fill=None, font=None,
alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def format_page(filename, name):

    template = load_workbook(r'''C:\Users\afrifa_k\Desktop\TEMPLATE 2019.xlsx''')
    ws_temp = template.active
    unformatted = load_workbook(filename)
    ws = unformatted.active

    ws.delete_rows(1)
    ws.delete_cols(1, 2)

    ws.page_margins.left = ws_temp.page_margins.left
    ws.page_margins.right = ws_temp.page_margins.right
    ws.page_margins.top = ws_temp.page_margins.top
    ws.page_margins.bottom = ws_temp.page_margins.bottom
    ws.page_margins.header = ws_temp.page_margins.header
    ws.page_margins.footer = ws_temp.page_margins.footer
    for col in ws.iter_cols(1, ws.max_column):
        ws.column_dimensions = ws_temp.column_dimensions
        ws.row_dimensions = ws_temp.row_dimensions
    ws_temp.delete_rows(2)
    for row in ws_temp.rows:
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.col_idx)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    ws['W6'].font = Font(name='Arial', size=9, bold=True, italic=True)
    ws.insert_rows(1)
    width = 49
    height = 35
    img = Image.open(r'''C:\Users\afrifa_k\Desktop\AAA-SOS-icon.jpg''')
    img = img.resize((width, height), Image.NEAREST)
    img.save(r'''C:\Users\afrifa_k\Desktop\AAA-SOS-icon.jpg''')
    img = openpyxl.drawing.image.Image(r'''C:\Users\afrifa_k\Desktop\AAA-SOS-icon.jpg''')
    ws.add_image(img, 'A1')
    ws.merge_cells('B7:E7')
    ws.merge_cells('F7:J7')
    ws.merge_cells('B8:E8')
    ws.merge_cells('B9:C9')
    ws.merge_cells('D9:E9')
    ws.merge_cells('F9:G9')
    ws.merge_cells('H9:J9')

    ws.merge_cells('M7:P7')
    ws.merge_cells('Q7:U7')
    ws.merge_cells('M8:P8')
    ws.merge_cells('F8:J8')
    ws.merge_cells('Q8:U8')

    ws.merge_cells('M9:N9')
    ws.merge_cells('O9:P9')
    ws.merge_cells('Q9:R9')
    ws.merge_cells('S9:U9')

    ws['A3'] = "NAME: " + str(name) + "          NATIONALITY :                SEX : MALE                DATE : MAY 2018"

    ws.oddHeader.center.text = "SOS-HERMANN GMEINER INTERNATIONAL COLLEGE                                                                                          TRANSCRIPT OF ACADEMIC RECORD"
    ws.oddHeader.center.size = 14
    ws.oddHeader.center.font = "Arial,Bold"
    ws.oddHeader.center.color = "000000"

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for range in ws.merged_cells.ranges:
        style_range(ws, str(range), border=border)

    ws.set_printer_settings(paper_size=9, orientation='landscape')
    ws.page_setup.fitToPage = True
    unformatted.save(filename)





